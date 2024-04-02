import numpy as np
import numba
import matplotlib.pyplot as plt
import math
import openpyxl

# some useful plot defaults
plt.rcParams.update({'font.size' : 20, 'lines.linewidth' : 3.5, 'figure.figsize' : (13,7)})

#Krusell Smith with aggregate variable in value function
amin,amax,na = -2, 3000, 201
a_space = np.linspace(amin,amax,na)
kmin,kman,nk = 140, 340, 6
k_space = np.linspace(kmin,kman,nk)
sin = 0.25
#income(w,b,tau)
#array([[0.5],
#       [1. ]])
eta=2
beta=0.995
tol=1E-7
alpha=0.36
delta=0.005
tau = 0
b = 0.000001
N = 0.5/(1.5-0.9565)
#Number of transition period
T=3000
Z=1

def rfn(Z,N,K,alpha,delta):
    return Z * alpha * (N/K)**(1-alpha) - delta

def wfn(Z,N,K,alpha):
    return Z * (1-alpha) * (K/N)**alpha

def income(Z,w,b,tau,N,K,alpha):
    w = wfn(Z,N,K,alpha)
    return np.array([[b],[(1-tau)*w]]) #[2,1] metrix, [e=u,e=e]

def Transition_metrics(p,q): 
    return np.array([[p,1-p],[q,1-q]])

def Stationary_distribution(p,q):
    n=Transition_metrics(p,q).shape[0]
    #initial guess of stationary distribution: uniform
    pi = np.full(n,1/n) #pi = [0.5,0.5]
    #Iterate untill stationary
    for i in range(1000):
        pi_next = Transition_metrics(p,q).T @ pi
        if np.max(np.abs(pi_next-pi)) < 1E-10:
            return pi
        else:
            pi = pi_next
            continue
        
p,q = 0.5, 0.0435 
Pi = Transition_metrics(p,q)
pi = Stationary_distribution(p,q)

def backward_iteration(Z,Va_init,Pi,w,b,tau,r,beta,eta,K):
    Va1_init = Pi @ Va_init #Expectation on Va next period with inital Va
    c_endog = (beta * Va1_init)**(-1/eta) #endogenous c with initial Va
    coh = income(Z,w,b,tau,N,K,alpha) + (1+(1-tau)*r) * a_space #cash on hand
    #Generating a' by interpolating c_endog+a' on cash one hand
    a1 = np.empty_like(coh)
    for e in range(2):
        a1[e,:]=np.interp(coh[e,:], c_endog[e, :] + a_space, a_space) #linear interp
    a1 = np.maximum(a1, a_space[0]) #limitation on borrowing
    c = coh - a1
    Va_new = (1+(1-tau)*r) * c**(-eta) #envelope condition
    return Va_new, a1, c
Hs=0
def steady_state_policy(Z,N,K,alpha,Pi,a_space,w,b,r,tau,eta,beta,tol):
    #initial guess of Va_init: assuming consume 5% of cash on hand
    coh_guess = income(Z,w,b,tau,N,K,alpha) + (1+(1-tau)*r) * a_space
    c_guess = 0.05 * coh_guess
    Va_guess = (1+(1-tau)*r) * c_guess**(-eta)
    Va = Va_guess
    #Iteration until error of new and old asset below tolerance:
    for _ in range(10000): #Maximum steps of iteration: 10000
        Va, a, c = backward_iteration(Z,Va,Pi,w,b,tau,r,beta,eta,K)
        if _>0 and np.max(np.abs(a - a_old)) < tol:
            if Hs==0:
                print('Backward iteration stops at step '+str(_)+'. Error: '+str(np.max(np.abs(a - a_old))))
            return Va, a, c
        else:
            a_old = a
            continue

@numba.njit
def forward_iteration(a_ss, a_space, f0):
    f1 = np.zeros_like(f0)
    for e in range(2):
        for i in range(len(a_space)):
            #Search for the interval of household [e,i]'s asset in the asseet grid space
            for n in range(len(a_space)):
                if a_ss[e,i] >= a_space[n] and a_ss[e,i] <= a_space[n+1]:
                    #print(e,i,a_ss[e,i],n)
                    interval = [n,n+1]
                    break
                else:
                    continue
            
            weight = (a_space[interval[1]] - a_ss[e,i])/(a_space[interval[1]] - a_space[interval[0]])
            f1[e,interval[0]] += weight * f0[e,i]
            f1[e,interval[1]] += (1-weight) * f0[e,i]
    return Pi.T @ f1

def steady_distribution(a_ss, a_space, tol):
    #initiate uniform distribution
    f0 = pi.reshape((2, 1)) * np.ones_like(a_space) / len(a_space)
    #initiate forward distribution
    for _ in range(1000000000):
        f1 = forward_iteration(a_ss, a_space, f0)
        if np.max(np.abs(f1-f0)) < tol:
            print('Forward iteration stops at step '+str(_)+'. Error: '+str(np.max(np.abs(f1-f0))))
            return f1
            break
        f0 = f1

def Kfn(f, a_ss):
    K = 0
    for e in range(2):
        for a in range(len(a_space)):
            K += f[e,a] * a_ss[e,a]
    return K


#Steady state distribution and parameters
from_xlsx = 1
if from_xlsx == 0:
    K_ss_1 = 200
    for _ in range(10000):
        w=wfn(Z,N,K_ss_1,alpha)
        r=rfn(Z,N,K_ss_1,alpha,delta)
        Va_ss, a_ss ,c_ss = steady_state_policy(Z,N,K_ss_1,alpha,Pi,a_space,w,b,r,tau,eta,beta,tol)
        f_ss = steady_distribution(a_ss, a_space, tol)
        K_ss = Kfn(f_ss, a_ss)
        if abs(K_ss-K_ss_1) < 1E-7: #Y - C - K * delta
            print('Market clears at step '+str(_)+'. Steady State K :')
            print(K_ss)
            break
        v = 0.005
        K_ss_1 = v * K_ss + (1-v) * K_ss_1
else:
    loca=r'...\note_MITshock2018.xlsx'
    wb=openpyxl.load_workbook(loca)
    Va_ss,f_ss,a_ss,c_ss,K_ss = np.zeros((2,201)),np.zeros((2,201)),np.zeros((2,201)),np.zeros((2,201)),250.5376092375545
    sheet=wb['f']
    for i in range(2):
        for n in range(201):
             f_ss[i][n] = sheet.cell(i+1,n+1).value
    sheet=wb['c']
    for i in range(2):
        for n in range(201):
             c_ss[i][n] = sheet.cell(i+1,n+1).value
    sheet=wb['a']
    for i in range(2):
        for n in range(201):
             a_ss[i][n] = sheet.cell(i+1,n+1).value
    sheet=wb['Va']
    for i in range(2):
        for n in range(201):
             Va_ss[i][n] = sheet.cell(i+1,n+1).value


#1. Guess T the economy goes back to steady state
T = 200
T_path = np.linspace(0,200,201)
#Shock: AR(1) TFP shock
Z_path = [1] + [1+1*0.01*0.9**(t+1) for t in range(T)]
Z_path = np.array(Z_path)

#2. Guess a path of K
K_path = np.linspace(K_ss,K_ss,T+1)
#compute r path and w path accordingly
r_path, w_path = np.zeros_like(K_path), np.zeros_like(K_path)
for t in range(len(K_path)):
    r_path[t] = rfn(Z_path[t],N,K_path[t],alpha,delta)
    w_path[t] = wfn(Z_path[t],N,K_path[t],alpha)

#3. Value function backwards iteration
Hs = 1
Va_path, a_path, c_path = np.zeros((T+1,a_ss.shape[0],a_ss.shape[1])), np.zeros((T+1,a_ss.shape[0],a_ss.shape[1])), np.zeros((T+1,a_ss.shape[0],a_ss.shape[1]))
Va_path[-1], a_path[-1], c_path[-1] = Va_ss, a_ss, c_ss
for t in reversed(range(T)):
    Va_path[t], a_path[t], c_path[t] = steady_state_policy(Z_path[t],N,K_path[t],alpha,Pi,a_space,w_path[t],b,r_path[t],tau,eta,beta,tol)

#4. Forward iteration of distribution function
f_path = np.zeros((T+1,a_ss.shape[0],a_ss.shape[1]))
f_path[0] = f_ss
for t in range(1,T+1):
    f_path[t] = forward_iteration(a_path[t], a_space, f_path[t-1])

#5. Compute new K path
K_path_new = []
for t in range(T+1):
    K_new = 0
    for e in range(2):
        K_new += np.vdot(f_path[t][e],a_space)
    K_path_new.append(K_new)
K_path_new = np.array(K_path_new)

#6. Compute the maximum error
error = np.max(np.abs(K_path_new - K_path))

pre = 0
if pre == 1:
    plt.plot(T_path,K_path,label = 'Initial guess path')
    plt.plot(T_path,K_path_new,label = 'first iteration')
    plt.ylabel('K')
    plt.xlabel('Period')
    plt.legend()
    plt.show()

#Interation till guess path of K converge
T = 300
T_path = np.linspace(0,T,T+1)
rho = 0.95
delta_z = 0.007
Z_path = [1] + [1+1*delta_z*rho**(t+1) for t in range(T)]
Z_path = np.array(Z_path)
K_path = np.linspace(K_ss,K_ss,T+1)
for _ in range(10000):
    r_path, w_path = np.zeros_like(K_path), np.zeros_like(K_path)
    for t in range(len(K_path)):
        r_path[t] = rfn(Z_path[t],N,K_path[t],alpha,delta)
        w_path[t] = wfn(Z_path[t],N,K_path[t],alpha)
    Va_path, a_path, c_path = np.zeros((T+1,a_ss.shape[0],a_ss.shape[1])), np.zeros((T+1,a_ss.shape[0],a_ss.shape[1])), np.zeros((T+1,a_ss.shape[0],a_ss.shape[1]))
    Va_path[-1], a_path[-1], c_path[-1] = Va_ss, a_ss, c_ss
    for t in reversed(range(T)):
        Va_path[t], a_path[t], c_path[t] = steady_state_policy(Z_path[t],N,K_path[t],alpha,Pi,a_space,w_path[t],b,r_path[t],tau,eta,beta,tol)
    f_path = np.zeros((T+1,a_ss.shape[0],a_ss.shape[1]))
    f_path[0] = f_ss
    for t in range(1,T+1):
        f_path[t] = forward_iteration(a_path[t], a_space, f_path[t-1])
    K_path_new = []
    for t in range(T+1):
        K_new = 0
        for e in range(2):
            K_new += np.vdot(f_path[t][e],a_space)
        K_path_new.append(K_new)
    K_path_new = np.array(K_path_new)
    error = np.max(np.abs(K_path_new - K_path))
    print(_,error)
    if error < 1E-5:
        break
    else:
        v = 0.8
        K_path = v * K_path_new + (1-v) * K_path

#Path found, compute the path of all aggregate variables, plot the impluse 
C_path, Y_path = np.linspace(0,T,T+1), np.linspace(0,T,T+1)
for t in range(T+1):
    C=0
    Y=0
    for e in range(2):
        C+=np.vdot(f_path[t][e],c_path[t][e])
    C_path[t] = C
    Y_path[t] = Z_path[t]*K_path[t]**alpha*N**(1-alpha)

def abs2precent(path):
    ss = path[0]
    return(path - ss)/ss


def plot(Z_path,w_path,K_path,C_path,r_path,Y_path):
    Z_path = abs2precent(Z_path)
    w_path = abs2precent(w_path)
    r_path = abs2precent(r_path)
    K_path = abs2precent(K_path)
    C_path = abs2precent(C_path)
    Y_path = abs2precent(Y_path)

    #plot
    Z_path_pct = Z_path * 100
    K_path_pct = K_path * 100
    w_path_pct = w_path * 100
    r_path_pct = r_path * 100
    C_path_pct = C_path * 100
    Y_path_pct = Y_path * 100

    # 创建一个2×3的子图矩阵
    fig, axes = plt.subplots(2, 3, figsize=(12, 8))

    # 绘制 Z_path 图
    axes[0, 0].plot(T_path[:200], Z_path_pct[:200])
    axes[0, 0].set_xlabel('Period')
    axes[0, 0].set_ylabel('%')
    axes[0, 0].set_title('Z')

    # 绘制 K_path 图
    axes[0, 1].plot(T_path[:200], K_path_pct[:200])
    axes[0, 1].set_xlabel('Period')
    axes[0, 1].set_ylabel('%')
    axes[0, 1].set_title('K')

    # 绘制 w_path 图
    axes[0, 2].plot(T_path[:200], w_path_pct[:200])
    axes[0, 2].set_xlabel('Period')
    axes[0, 2].set_ylabel('%')
    axes[0, 2].set_title('w')

    # 绘制 r_path 图
    axes[1, 0].plot(T_path[:200], r_path_pct[:200])
    axes[1, 0].set_xlabel('Period')
    axes[1, 0].set_ylabel('%')
    axes[1, 0].set_title('r')

    # 绘制 C_path 图
    axes[1, 1].plot(T_path[:200], C_path_pct[:200])
    axes[1, 1].set_xlabel('Period')
    axes[1, 1].set_ylabel('%')
    axes[1, 1].set_title('C')

    # 绘制 Y_path 图
    axes[1, 2].plot(T_path[:200], Y_path_pct[:200])
    axes[1, 2].set_xlabel('Period')
    axes[1, 2].set_ylabel('%')
    axes[1, 2].set_title('Y')

    # 调整子图之间的间距
    plt.tight_layout()

    # 显示图表
    plt.show()

pre = 1
if pre == 1:
    plot(Z_path,w_path,K_path,C_path,r_path,Y_path)

