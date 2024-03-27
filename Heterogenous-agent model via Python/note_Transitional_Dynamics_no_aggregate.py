import numpy as np
import numba
from scipy import optimize
import matplotlib.pyplot as plt
import statsmodels.api as sm
import math
import openpyxl

# some useful plot defaults
plt.rcParams.update({'font.size' : 20, 'lines.linewidth' : 3.5, 'figure.figsize' : (13,7)})

#Krusell Smith with aggregate variable in value function
amin,amax,na = -2, 3000, 201
a_space = np.linspace(amin,amax,na)
kmin,kman,nk = 140, 340, 6
k_space = np.linspace(kmin,kman,nk)
sin = 0.25
#income(w,b,tau)
#array([[0.5],
#       [1. ]])
eta=2
beta=0.995
tol=1E-7
alpha=0.36
delta=0.005
tau = 0
b = 0.000001
N = 0.5/(1.5-0.9565)
#Number of transition period
T=3000

def rfn(N,K,alpha,delta):
    return alpha * (N/K)**(1-alpha) - delta

def wfn(N,K,alpha):
    return (1-alpha) * (K/N)**alpha

def income(w,b,tau,N,K,alpha):
    w = wfn(N,K,alpha)
    return np.array([[b],[(1-tau)*w]]) #[2,1] metrix, [e=u,e=e]

def Transition_metrics(p,q): 
    return np.array([[p,1-p],[q,1-q]])

def Stationary_distribution(p,q):
    n=Transition_metrics(p,q).shape[0]
    #initial guess of stationary distribution: uniform
    pi = np.full(n,1/n) #pi = [0.5,0.5]
    #Iterate untill stationary
    for i in range(1000):
        pi_next = Transition_metrics(p,q).T @ pi
        if np.max(np.abs(pi_next-pi)) < 1E-10:
            return pi
        else:
            pi = pi_next
            continue
        
p,q = 0.5, 0.0435 
Pi = Transition_metrics(p,q)
pi = Stationary_distribution(p,q)

def backward_iteration(Va_init,Pi,w,b,tau,r,beta,eta,K):
    Va1_init = Pi @ Va_init #Expectation on Va next period with inital Va
    c_endog = (beta * Va1_init)**(-1/eta) #endogenous c with initial Va
    coh = income(w,b,tau,N,K,alpha) + (1+(1-tau)*r) * a_space #cash on hand
    #Generating a' by interpolating c_endog+a' on cash one hand
    a1 = np.empty_like(coh)
    for e in range(2):
        a1[e,:]=np.interp(coh[e,:], c_endog[e, :] + a_space, a_space) #linear interp
    a1 = np.maximum(a1, a_space[0]) #limitation on borrowing
    c = coh - a1
    Va_new = (1+(1-tau)*r) * c**(-eta) #envelope condition
    return Va_new, a1, c
Hs=0
def steady_state_policy(N,K,alpha,Pi,a_space,w,b,r,tau,eta,beta,tol):
    #initial guess of Va_init: assuming consume 5% of cash on hand
    coh_guess = income(w,b,tau,N,K,alpha) + (1+(1-tau)*r) * a_space
    c_guess = 0.05 * coh_guess
    Va_guess = (1+(1-tau)*r) * c_guess**(-eta)
    Va = Va_guess
    #Iteration until error of new and old asset below tolerance:
    for _ in range(10000): #Maximum steps of iteration: 10000
        Va, a, c = backward_iteration(Va,Pi,w,b,tau,r,beta,eta,K)
        if _>0 and np.max(np.abs(a - a_old)) < tol:
            if Hs==0:
                print('Backward iteration stops at step '+str(_)+'. Error: '+str(np.max(np.abs(a - a_old))))
            return Va, a, c
        else:
            a_old = a
            continue

@numba.njit
def forward_iteration(a_ss, a_space, f0):
    f1 = np.zeros_like(f0)
    for e in range(2):
        for i in range(len(a_space)):
            #Search for the interval of household [e,i]'s asset in the asseet grid space
            for n in range(len(a_space)):
                if a_ss[e,i] >= a_space[n] and a_ss[e,i] <= a_space[n+1]:
                    #print(e,i,a_ss[e,i],n)
                    interval = [n,n+1]
                    break
                else:
                    continue
            
            weight = (a_space[interval[1]] - a_ss[e,i])/(a_space[interval[1]] - a_space[interval[0]])
            f1[e,interval[0]] += weight * f0[e,i]
            f1[e,interval[1]] += (1-weight) * f0[e,i]
    return Pi.T @ f1

def steady_distribution(a_ss, a_space, tol):
    #initiate uniform distribution
    f0 = pi.reshape((2, 1)) * np.ones_like(a_space) / len(a_space)
    #initiate forward distribution
    for _ in range(1000000000):
        f1 = forward_iteration(a_ss, a_space, f0)
        if np.max(np.abs(f1-f0)) < tol:
            print('Forward iteration stops at step '+str(_)+'. Error: '+str(np.max(np.abs(f1-f0))))
            return f1
            break
        f0 = f1

def Kfn(f, a_ss):
    K = 0
    for e in range(2):
        for a in range(len(a_space)):
            K += f[e,a] * a_ss[e,a]
    return K

n_300 = 0
for _ in range(len(a_space)):
    if a_space[_] < 300:
        n_300 = _

#Initial distribution
shape = np.zeros((T,2,na)) #state: (T,e,a)
f_init = np.zeros((2,na))
for e in range(2):
    for a in range(na):
        if a <= n_300:
            f_init[e,a] = 1/2/(n_300+1)
        else:
            f_init[e,a] = 0

K_init = 0
for e in range(2):
    K_init += np.vdot(f_init[e,:],a_space)
loca=r'D:\onedrive\OneDrive - City University of Hong Kong - Student\大学博一上主课\nber-workshop-2023-main\2024computation_note\note_TD_no_aggregate.xlsx'
#Change loca to your own path

wb=openpyxl.load_workbook(loca)
sheet=wb['f_ss']
f_ss = np.zeros((2,201))
for i in range(1,3):
    for n in range(1,202):
        f_ss[i-1][n-1] = sheet.cell(i,n).value

sheet=wb['Va_ss']
Va_ss = np.zeros((2,201))
for i in range(1,3):
    for n in range(1,202):
        Va_ss[i-1][n-1] = sheet.cell(i,n).value

sheet=wb['a_ss']
a_ss = np.zeros((2,201))
for i in range(1,3):
    for n in range(1,202):
        a_ss[i-1][n-1] = sheet.cell(i,n).value

sheet=wb['c_ss']
c_ss = np.zeros((2,201))
for i in range(1,3):
    for n in range(1,202):
        c_ss[i-1][n-1] = sheet.cell(i,n).value

K_ss = 0
for e in range(2):
    K_ss += np.vdot(f_ss[e],a_ss[e])

#Guess the time path of K, r, w
T = 500
K_guess = np.linspace(K_init,K_ss,T)
r_guess = np.zeros_like(K_guess)
w_guess = np.zeros_like(K_guess)
for k in range(len(K_guess)):
    r_guess[k] = rfn(N,K_guess[k],alpha,delta)
    w_guess[k] = wfn(N,K_guess[k],alpha)
shape = [T,2,na]
Hs = 1

for _ in range(10000):
    print('Iteration step: '+str(_))
    f = f_init
    flist = []
    flist.append(f)
    Va = np.zeros(shape)
    a = np.zeros(shape)
    c = np.zeros(shape)
    Va[-1] = Va_ss
    a[-1] = a_ss
    c[-1] = c_ss
    for t in reversed(range(T-1)): #Backward iteration from T-1 to 1
        w = w_guess[t]
        r = r_guess[t]
        Va[t], a[t] ,c[t] = steady_state_policy(N,K_guess[t],alpha,Pi,a_space,w,b,r,tau,eta,beta,tol)
    for t in range(1,T):
        f = forward_iteration(a[t],a_space,flist[-1])
        flist.append(f)
    flist = np.array(flist)
    #Compute new path of K, r and w
    K_guess_new = np.zeros_like(K_guess) 
    for t in range(T):
        for e in range(2):
            K_guess_new[t] += np.vdot(flist[t,e],a[t,e,:])
    if _ ==1  or _ == 0 or _ == 3 or _ == 5  or _ == 10:
        plt.plot(np.linspace(1,1+T,T),K_guess_new,label = 'Iteration Step: '+str(_))
    if np.max(np.abs(K_guess_new - K_guess)) < tol:
        print(_,np.max(np.abs(K_guess_new - K_guess)))
        plt.plot(np.linspace(1,1+T,T),K_guess_new,label = 'Converged K Path')
        plt.legend()
        plt.show()
        break
    else:
        v = 0.7
        K_guess = v * K_guess_new + (1-v) * K_guess
        for k in range(len(K_guess)):
            r_guess[k] = rfn(N,K_guess[k],alpha,delta)
            w_guess[k] = wfn(N,K_guess[k],alpha)
    print('Maximum error :')
    print(np.max(np.abs(K_guess_new - K_guess)))






